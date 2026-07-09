from __future__ import annotations

import json
from argparse import ArgumentParser, Namespace
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install it with: python -m pip install openpyxl"
    ) from exc


SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parents[1]
DEFAULT_METRIC = "expected_fidelity"
DEFAULT_INPUT_DIR = SCRIPT_DIR
DEFAULT_OUTPUT_DIR = ROOT / "output/spreadsheets"

CORE_FEATURE_NAMES = [
    "num_qubits",
    "depth",
    "program_communication",
    "critical_depth",
    "entanglement_ratio",
    "parallelism",
    "liveness",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
SUMMARY_FILL = PatternFill("solid", fgColor="F2F6FA")
THIN_BLUE = Side(style="thin", color="D9E2F3")
THIN_GREY = Side(style="thin", color="E7EAF0")


def feature_description(name: str) -> tuple[str, str]:
    if name.startswith("gate_count_"):
        gate = name.removeprefix("gate_count_")
        return "Gate count", f"Conteggio del gate OpenQASM '{gate}' nel circuito sorgente."

    descriptions = {
        "num_qubits": ("Basic circuit size", "Numero di qubit logici del circuito sorgente."),
        "depth": (
            "Basic circuit size",
            "Profondita' del circuito sorgente prima della compilazione hardware-specific.",
        ),
        "program_communication": (
            "Structural feature",
            "Misura SupermarQ legata alle interazioni/comunicazioni tra qubit.",
        ),
        "critical_depth": (
            "Structural feature",
            "Frazione di gate a due qubit sul percorso critico; valori alti indicano maggiore sequenzialita'.",
        ),
        "entanglement_ratio": (
            "Structural feature",
            "Rapporto tra operazioni entangling/a due qubit e operazioni totali.",
        ),
        "parallelism": ("Structural feature", "Misura della possibilita' di eseguire operazioni in parallelo."),
        "liveness": ("Structural feature", "Frazione della matrice qubit-tempo in cui i qubit sono attivi."),
    }
    return descriptions.get(name, ("Other", ""))


def add_table(ws: Any, name: str) -> None:
    ref = ws.dimensions
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(ws: Any) -> None:
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(left=THIN_BLUE, right=THIN_BLUE, top=THIN_BLUE, bottom=THIN_BLUE)


def style_body(ws: Any) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
            cell.alignment = Alignment(vertical="top")


def write_rows(ws: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)


def apply_number_format(ws: Any, start_col: int, end_col: int, start_row: int, end_row: int, fmt: str) -> None:
    if end_col < start_col or end_row < start_row:
        return
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.number_format = fmt


def feature_format(feature_name: str) -> str:
    if feature_name.startswith("gate_count_") or feature_name in {"num_qubits", "depth"}:
        return "0"
    return "0.000000"


def set_feature_formats(ws: Any, feature_names: list[str], first_col: int) -> None:
    for offset, feature_name in enumerate(feature_names):
        col_idx = first_col + offset
        apply_number_format(ws, col_idx, col_idx, 2, ws.max_row, feature_format(feature_name))


def set_reasonable_widths(ws: Any, min_width: int = 10, max_width: int = 36) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(1, min(ws.max_row, 20) + 1)]
        longest = max((len(str(value)) for value in values if value is not None), default=min_width)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def score_columns(payload: dict[str, Any]) -> list[str]:
    configured = payload.get("score_columns") or []
    if configured:
        return [str(column) for column in configured]
    max_score_count = max((len(row.get("score_values") or []) for row in payload.get("rows", [])), default=0)
    if max_score_count <= 1:
        return ["score"]
    return [f"score_{idx}" for idx in range(max_score_count)]


def padded_scores(row: dict[str, Any], width: int) -> list[float | None]:
    values = list(row.get("score_values") or [])
    return [*values, *([None] * max(0, width - len(values)))]


def joined_values(values: Any) -> str:
    if not values:
        return ""
    return ", ".join(str(value) for value in values)


def ordered_dataset_features(payload: dict[str, Any]) -> list[str]:
    gate_feature_names = [feature for feature in payload["feature_names"] if feature.startswith("gate_count_")]
    ordered_core_feature_names = [feature for feature in CORE_FEATURE_NAMES if feature in payload["feature_names"]]
    other_feature_names = [
        feature
        for feature in payload["feature_names"]
        if feature not in ordered_core_feature_names and feature not in gate_feature_names
    ]
    return ordered_core_feature_names + other_feature_names + gate_feature_names


def default_input_path(metric: str, input_dir: Path) -> Path:
    return input_dir / f"device_selector_dataset_{metric}.json"


def default_output_path(metric: str, output_dir: Path) -> Path:
    return output_dir / f"MQT_device_selector_dataset_{metric}.xlsx"


def build_workbook(input_path: Path, output_path: Path | None = None) -> Path:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    metric = str(payload["metric"])
    output = output_path or default_output_path(metric, DEFAULT_OUTPUT_DIR)
    if not payload.get("rows"):
        raise SystemExit(f"Input dataset has no rows: {input_path}")

    dataset_feature_names = ordered_dataset_features(payload)
    score_headers = score_columns(payload)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    dataset = wb.create_sheet("Dataset")
    xmatrix = wb.create_sheet("X Matrix")
    legend = wb.create_sheet("Feature Legend")

    id_headers = [
        "index",
        "name",
        "label_device",
        *score_headers,
        "qasm_found",
        "qasm_path",
        "source_num_qubits",
        "source_depth",
        "compiled_qasm_count",
        "compiled_qasm_devices",
    ]
    dataset_headers = id_headers + dataset_feature_names
    dataset_rows = [
        [
            row["index"],
            row["name"],
            row["label_device"],
            *padded_scores(row, len(score_headers)),
            row.get("qasm_found"),
            row.get("qasm_path"),
            row.get("source_num_qubits"),
            row.get("source_depth"),
            row.get("compiled_qasm_count"),
            joined_values(row.get("compiled_qasm_devices")),
            *[row["features"].get(feature) for feature in dataset_feature_names],
        ]
        for row in payload["rows"]
    ]
    write_rows(dataset, [dataset_headers, *dataset_rows])
    add_table(dataset, "DeviceSelectorDataset")
    dataset.sheet_view.showGridLines = False
    dataset.column_dimensions["A"].width = 8
    dataset.column_dimensions["B"].width = 28
    dataset.column_dimensions["C"].width = 22
    for col_idx in range(4, dataset.max_column + 1):
        dataset.column_dimensions[get_column_letter(col_idx)].width = 14
    style_header(dataset)
    style_body(dataset)
    score_start_col = 4
    score_end_col = score_start_col + len(score_headers) - 1
    apply_number_format(dataset, score_start_col, score_end_col, 2, dataset.max_row, "0.0000000000")
    metadata_width = 6
    set_feature_formats(dataset, dataset_feature_names, score_end_col + metadata_width + 1)

    x_headers = ["name", *payload["feature_names"]]
    x_rows = [
        [row["name"], *[row["features"].get(feature) for feature in payload["feature_names"]]]
        for row in payload["rows"]
    ]
    write_rows(xmatrix, [x_headers, *x_rows])
    add_table(xmatrix, "XMatrix")
    xmatrix.sheet_view.showGridLines = False
    xmatrix.column_dimensions["A"].width = 28
    for col_idx in range(2, xmatrix.max_column + 1):
        xmatrix.column_dimensions[get_column_letter(col_idx)].width = 14
    style_header(xmatrix)
    style_body(xmatrix)
    set_feature_formats(xmatrix, payload["feature_names"], 2)

    legend_rows = [
        ["feature_name", "group", "description"],
        *[[feature, *feature_description(feature)] for feature in payload["feature_names"]],
    ]
    write_rows(legend, legend_rows)
    add_table(legend, "FeatureLegend")
    legend.sheet_view.showGridLines = False
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 22
    legend.column_dimensions["C"].width = 80
    style_header(legend)
    style_body(legend)
    for row in legend.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    label_names = list((payload.get("label_distribution") or {}).keys())
    if len(label_names) == 1:
        candidate_summary = (
            "In questo dataset locale c'e' un solo device candidato, "
            f"quindi tutte le label sono {label_names[0]}."
        )
    else:
        candidate_summary = (
            f"In questo dataset locale ci sono {len(label_names)} device candidati: "
            f"{', '.join(label_names)}."
        )

    summary_rows = [
        ["Campo", "Valore"],
        ["Figure of merit", metric],
        ["Numero campioni", payload["sample_count"]],
        ["Numero feature in X", payload["feature_count"]],
        ["Foglio Dataset", "Vista leggibile: metadati, score e feature principali, con gate count spostati a destra."],
        ["Foglio X Matrix", "Nome circuito + tutte le colonne della matrice X nell'ordine originale esportato da MQT."],
        ["File sorgenti", f"training_data_{metric}.npy + names_list_{metric}.npy + scores_list_{metric}.npy"],
        ["Interpretazione X", "Feature vector del circuito target-independent, non del circuito compilato."],
        ["Interpretazione y", "Device migliore secondo lo score della figure of merit."],
        ["Nota dataset", candidate_summary],
        ["Distribuzione label", json.dumps(payload["label_distribution"])],
    ]
    write_rows(summary, summary_rows)
    summary.sheet_view.showGridLines = False
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 96
    style_header(summary)
    for cell in summary["A"][1:]:
        cell.fill = SUMMARY_FILL
        cell.font = Font(name="Calibri", size=10, bold=True)
    style_body(summary)
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    return output


def parse_args() -> Namespace:
    parser = ArgumentParser(description="Build a readable Excel workbook from a device selector dataset JSON.")
    parser.add_argument("--metric", default=DEFAULT_METRIC, help="Figure of merit to build when --input is not passed.")
    parser.add_argument("--input", type=Path, help="Input JSON path. Defaults to device_selector_dataset_<metric>.json.")
    parser.add_argument("--output", type=Path, help="Output XLSX path. Defaults to MQT_device_selector_dataset_<metric>.xlsx.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="Directory containing generated JSON files.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated XLSX files.")
    parser.add_argument(
        "--all-available",
        action="store_true",
        help="Build one workbook for every device_selector_dataset_*.json file in --input-dir.",
    )
    return parser.parse_args()


def metric_from_input_path(path: Path) -> str:
    stem = path.stem
    return stem.removeprefix("device_selector_dataset_")


def main() -> None:
    args = parse_args()
    if args.all_available:
        input_paths = sorted(args.input_dir.glob("device_selector_dataset_*.json"))
        if not input_paths:
            raise SystemExit(f"No device_selector_dataset_*.json files found in {args.input_dir}")
        for input_path in input_paths:
            metric = metric_from_input_path(input_path)
            build_workbook(input_path, default_output_path(metric, args.output_dir))
        return

    input_path = args.input or default_input_path(args.metric, args.input_dir)
    output_path = args.output or default_output_path(args.metric, args.output_dir)
    build_workbook(input_path, output_path)


if __name__ == "__main__":
    main()
